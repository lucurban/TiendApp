'''
===============================================================================
                                  TiendaApp
===============================================================================

Descripción:
    Este programa registra ventas y actualiza el inventario de una tienda.
    Permite consultar existencias, modificar precios y generar un resumen de
    ventas para una fecha especificada por el usuario.

Autor: Lucas O. Urbano Bedoya
Fecha: Octubre 2025
Lenguaje: Python
Librerías: Pandas y Openpyxl
===============================================================================

===============================================================================
                          TiendaApp - Clase tienda
===============================================================================

Descripción:
   Esta clase gestiona el inventario y las ventas de una tienda.
   Utiliza Pandas y Openpyxl para leer y escribir en archivos Excel.
===============================================================================
'''

# --- Importar paquetes ---
import pandas as pd
from openpyxl import load_workbook

'''
===============================================================================
                        Definición de la clase tienda
===============================================================================
'''
class tienda:
    # --- Inicializador de la clase ---
    def __init__(self, inventario, producto=''):
        '''
        Constructor de la clase tienda.

        Parámetros:
        - inventario (str): Nombre del archivo Excel que contiene los datos.
        - producto (str, opcional): Nombre del producto a gestionar.
        '''
        self.inventario = inventario
        self.producto = producto
    
    '''
    ===========================================================================
                     Métodos principales de la clase
    ===========================================================================
    '''

    # --- Método de lectura del inventario ---
    def leer_inventario(self):
        '''
        Verifica si un producto se encuentra en el archivo de inventario.

        Retorna:
        - True si el producto existe.
        - False si no existe o si el archivo no se encuentra.
        '''
        try:
            # --- Leer el archivo Excel del inventario ---
            df_inventario = pd.read_excel(self.inventario)

            # --- Buscar si el producto existe ---
            for index, row in df_inventario.iterrows():
                # --- Si se encuentra el producto, retornar True ---
                if row['Producto'] == self.producto:
                   return True
            
            # --- Si no se encuentra el producto, retornar False ---
            return False
                
        except FileNotFoundError:
            # --- Si el archivo no existe, retornar False ---
            return False
        
    # --- Metodo para revisar el inventario ---
    def revisar(self):
        '''
        Muestra el contenido completo del inventario en pantalla.

        Lee el archivo Excel asociado al inventario y lo imprime como un 
        DataFrame. Si el archivo no existe, informa al usuario que aún 
        no se ha creado un inventario.

        Excepciones:
            FileNotFoundError: Se lanza si el archivo de inventario no existe.
        '''
        try:
            # Cargar el archivo de inventario en un DataFrame
            df_inventario = pd.read_excel(self.inventario)

            # Mostrar el inventario en consola
            print('\n', df_inventario)

        except FileNotFoundError:
            # Mensaje si el inventario aún no ha sido creado
            print('\nAun no se ha creado un inventario')

    # --- Metodo para actualizar el inventario ---
    def actualizar_inventario(self, producto):
        '''
        Actualiza la cantidad disponible de un producto en el inventario. 
        
        Parámetros:
            producto (str): Nombre del producto a actualizar. 
            
        Descripción: 
            - Busca el producto en la hoja 'Inventario' del archivo Excel. 
            - Si el producto no se encuentra, muestra un mensaje de aviso. 
            - Si se encuentra exactamente una coincidencia: 
                * Muestra la cantidad actual. 
                * Solicita al usuario la nueva cantidad (en unidades o gramos). 
                * Actualiza el valor en el DataFrame y lo sobrescribe en Excel. 
            - Si hay múltiples coincidencias, muestra un mensaje de error.
        '''
        # Leer el inventario desde el archivo Excel
        df_inventario = pd.read_excel(self.inventario, sheet_name='Inventario')
        contador = 0

        # Buscar el producto en el DataFrame
        for index, row in df_inventario.iterrows():
            if row['Producto'] == producto:
                contador += 1
                fila = index # Guarda la posición de la coincidencia

        # Evaluar los casos según la cantidad de coincidencias
        if contador < 1:
            print(f'{producto} no se encuentra en el inventario')

        elif contador == 1:
            print(f'{producto} se encuentra en el inventario')
            print(f'Actualmente hay {df_inventario.at[fila, "Cantidad"]} {df_inventario.at[fila, "Unidad"]} de {producto}')

            # Solicitar nueva cantidad según la unidad del producto
            if df_inventario.at[fila, 'Unidad'] == 'unidades':
                nueva_cant = int(input(f'¿Cuántas unidades de {producto} deben haber en el inventario? '))

            elif df_inventario.at[fila, 'Unidad'] == 'gramos':
                nueva_cant = int(input(f'¿Cuántos gramos de {producto} deben haber en el inventario? '))
            
            # Actualizar el valor en el DataFrame
            df_inventario.at[fila, 'Cantidad'] = nueva_cant

            # Escribir los cambios en el archivo Excel
            with pd.ExcelWriter(self.inventario,
                                mode='a',
                                if_sheet_exists='overlay'
                                ) as writer:
                 df_inventario.to_excel(writer,
                                   sheet_name='Inventario', 
                                   index=False, 
                                   header=True
                                   )

            print(f'El inventario de {producto} se actualizo con exito')
            print(f'Ahora hay {nueva_cant} {df_inventario.at[fila, "Unidad"]} de {producto} en el inventario')

        else:
            print(f'Error: Hay {contador} registros de {producto} en el inventario')

    # --- Metodo para agregar un producto al inventario ---
    def agregar(self):
        '''
        Agrega un nuevo producto al inventario y a la hoja de precios.
        
        Si el archivo de inventario ya existe, añade el nuevo registro a las
        hojas 'Inventario' y 'Precios'. Si no existe, crea el archivo y las 
        hojas con encabezados. También solicita la información necesaria al 
        usuario (cantidad, unidad, precios de compra y venta). 
        
        Retorna: 
            None
        '''
        # --- Solicitar información del nuevo producto ---
        cantidad = int(input(f'Ingresa la cantidad de {self.producto} que debe ingresarse al inventario: '))
        und_correcta = False

        # Validar unidad (solo acepta 'gramos' o 'unidades')
        while und_correcta == False:
                unidad = input(f'{self.producto} se mide en gramos o unidades: ')

                if (unidad == 'gramos') or (unidad == 'unidades'):
                    und_correcta = True

                else:
                    print('La unidad especificada no es correcta')

        # --- Calcular utilidad ---
        precio_compra = int(input('Ingresa el precio de compra: '))
        precio_venta = int(input('Ingresa el precio de venta: '))
        utilidad = precio_venta - precio_compra

        # --- Crear DataFrames con los nuevos registros ---
        df_nuevo_producto = pd.DataFrame({'Producto': [self.producto],
                                       'Cantidad': [cantidad],
                                       'Unidad': [unidad]
                                       }) 

        df_precios_np = pd.DataFrame({'Producto': [self.producto],
                                   'Precio Compra': [precio_compra],
                                   'Precio Venta': [precio_venta],
                                   'Utilidad': [utilidad]
                                   })

        # --- Agregar o crear el archivo de inventario ---
        try:
            # Si el archivo ya existe, lo abrimos en modo append (añadir)
            df_inventario = pd.read_excel(self.inventario, sheet_name='Inventario')
            df_precios = pd.read_excel(self.inventario, sheet_name='Precios')
            cant_productos = len(df_inventario)
            cant_prod_precios = len(df_precios)

            # Escribir nuevas filas sin sobrescribir encabezados
            with pd.ExcelWriter(self.inventario, 
                                mode='a',
                                if_sheet_exists='overlay'
                                ) as writer:
                df_nuevo_producto.to_excel(writer,
                                           sheet_name='Inventario',
                                           startrow=cant_productos+1,
                                           index=False,
                                           header=False
                                           )
                
                df_precios_np.to_excel(writer,
                                       sheet_name='Precios',
                                       startrow=cant_prod_precios+1,
                                       index=False,
                                       header=False
                                       )

            print(f'{self.producto} se agrego al inventario con exito')
        
        except FileNotFoundError:
            # Si el archivo no existe, se crea uno nuevo con ambas hojas
            with pd.ExcelWriter(self.inventario,
                                mode='w') as writer:
                df_nuevo_producto.to_excel(writer,
                                           sheet_name='Inventario',
                                           index=False
                                           )
                
                df_precios_np.to_excel(writer,
                                       sheet_name='Precios',
                                       index=False
                                       )

            print(f'{self.producto} se agrego al inventario con exito')
        
    # --- Metodo para actualizar los precios del inventario ---
    def actualizar_precios(self):
        '''
        Permite actualizar los precios de compra y venta de los productos del
        inventario. 
        
        Muestra la hoja de precios y permite al usuario seleccionar un producto 
        para modificar sus valores. Calcula automáticamente la nueva utilidad y 
        guarda los cambios en la hoja 'Precios' del archivo Excel.
        
        Retorna: 
            None
        '''
        # --- Leer hoja de precios ---
        df_precios = pd.read_excel(self.inventario, sheet_name='Precios')
        contador = 0
        act_precio = 's'

        print('\n', df_precios)

        # --- Bucle principal de actualización ---
        while act_precio == 's':
            act_precio = input('\n¿Deseas actualizar los precios de algún producto? s|n: ')
            
            if act_precio == 's':
                producto = input('¿De que producto deseas actualizar sus precios? ')

                # Buscar producto en la hoja de precios
                for index, row in df_precios.iterrows():
                    if row['Producto'] == producto:
                        contador += 1
                        fila = index

                # --- Validaciones según cantidad de coincidencias ---
                if contador < 1:
                    print(f'{producto} no se encuentra en el inventario')

                elif contador == 1:
                    print(f'\n{producto} se encuentra en el inventario')
                    print(f'Actualmente el precio de compra es {df_precios.at[fila, "Precio Compra"]}')
                    print(f'Actualmente el precio de venta es {df_precios.at[fila, "Precio Venta"]}')
                    print(f'Actualmente la utilidad es {df_precios.at[fila, "Utilidad"]}')

                    # Solicitar nuevos precios
                    nuevo_precio_compra = int(input(f'\n¿Cual debe ser el nuevo precio de compra de {producto}? '))
                    nuevo_precio_venta = int(input(f'¿Cual debe ser el nuevo precio de venta de {producto}? '))
                    nueva_utilidad = nuevo_precio_venta - nuevo_precio_compra

                    # Actualizar valores en el DataFrame
                    df_precios.at[fila, 'Precio Compra'] = nuevo_precio_compra
                    df_precios.at[fila, 'Precio Venta'] = nuevo_precio_venta
                    df_precios.at[fila, 'Utilidad'] = nueva_utilidad

                    # Guardar cambios en el archivo Excel
                    with pd.ExcelWriter(self.inventario,
                                        mode='a',
                                        if_sheet_exists='overlay'
                                        ) as writer:
                        df_precios.to_excel(writer,
                                        sheet_name='Precios', 
                                        index=False, 
                                        header=True
                                        )

                    print(f'\nLos precios de {producto} se actualizaron con exito')
                    print(f'Ahora el precio de compra es {nuevo_precio_compra}')
                    print(f'Ahora el precio de venta es {nuevo_precio_venta}')
                    print(f'Ahora la utilidad es {nueva_utilidad}')

                else:
                    print(f'Error: Hay {contador} registros de {self.producto} en el inventario')

            elif act_precio == 'n':
                print('\nLos precios ya fueron actualizados')

            else:
                print('La opción ingresada no es valida')
                act_precio = 's'

    # --- Método para registrar ventas de productos ---
    def vender(self):

        '''
        Registra una o varias ventas de productos, verificando disponibilidad 
        en el inventario. 
        
        El método permite: 
            - Validar si un producto existe en el inventario. 
            - Consultar el precio de venta y la unidad del producto. 
            - Ingresar la cantidad vendida y calcular el subtotal. 
            - Permitir al usuario agregar más productos a la misma venta. 
            - Guardar la venta en la hoja 'Ventas' del archivo Excel (creándola
              si no existe).
            
            Retorna: 
                DataFrame: con el detalle de la venta (productos, cantidades y
                subtotales).
        '''
        # --- Crear DataFrame para registrar los productos vendidos ---
        df_venta = pd.DataFrame(columns=['Fecha', 'Producto', 'Cantidad', 'Unidad', 'Precio', 'Subtotal'])
        
        # --- Leer inventario y precios desde el archivo Excel ---
        df_inventario = pd.read_excel(self.inventario, sheet_name='Inventario')
        df_precios = pd.read_excel(self.inventario, sheet_name='Precios')
        
        nuevo_prod = self.producto
        agregar_producto = True

        # --- Bucle principal: permite agregar varios productos a la venta ---
        while agregar_producto:
            prod_en_inventario = True
            contador = 0
            
            # --- Buscar el producto en el inventario ---
            for index, row in df_inventario.iterrows():
                if row['Producto'] == nuevo_prod:
                    fecha = pd.Timestamp.today().date()
                    cant_inv = row['Cantidad']
                    und = row['Unidad']
                    contador += 1

                # --- Validar existencia del producto ---
                if (contador < 1) or (contador > 1):
                    prod_en_inventario = False

                else:
                    prod_en_inventario = True

            # --- Si el producto no está en el inventario ---
            if prod_en_inventario == False:  
                print(f'Lo siento, no tengo {nuevo_prod}')
                print('Estos son los productos que te puedo ofrecer:')
                    
                # Mostrar productos disponibles
                for index, row in df_inventario.iterrows():
                    print(f'- {row['Producto']} ({row['Cantidad']} {row['Unidad']})')

                # Preguntar si desea otro producto
                resp_correcta = False

                while resp_correcta == False:
                    otro_prod = input('\n¿Deseas algo mas? s|n: ')
                    
                    if otro_prod == 's':
                        nuevo_prod = input('¿Que producto deseas?: ')
                        
                        for index, row in df_inventario.iterrows():
                            if row['Producto'] == nuevo_prod:
                                fecha = pd.Timestamp.today().date()
                                cant_inv = row['Cantidad']
                                und = row['Unidad']

                        resp_correcta = True
                        agregar_producto = True
                    
                    elif otro_prod == 'n':
                        print('Vuelve pronto!')
                        
                        resp_correcta = True
                        agregar_producto = False
                                                
                        return df_venta
                    
                    else:
                        print('La opción ingresada no es valida')
                        
                        resp_correcta = False
            
            # --- Buscar el precio del producto ---
            for index, row in df_precios.iterrows():
                if row['Producto'] == nuevo_prod:
                    precio = row['Precio Venta']

            # --- Solicitar cantidad deseada ---
            if cant_inv > 0:
                if und == 'unidades':
                    cant = int(input(f'cuantas {und} de {nuevo_prod} deseas: '))

                elif und == 'gramos':
                    cant = int(input(f'cuantos {und} de {nuevo_prod} deseas: '))

            # --- Calcular subtotal y agregar al DataFrame de venta ---
            subtotal = cant * precio

            df_venta.loc[len(df_venta)] = [fecha, nuevo_prod, cant, und, precio, subtotal]

            total = df_venta['Subtotal'].sum()

            # --- Confirmar si desea seguir comprando ---
            resp_correcta = False

            while resp_correcta == False:
                print('\n', df_venta)
                print(f'Total a pagar: {total}')

                otro_prod = input('\n¿Deseas algo mas? s|n: ')

                if otro_prod == 's':
                    nuevo_prod = input('¿Que producto deseas?: ')
                    
                    agregar_producto = True
                    resp_correcta = True

                elif otro_prod == 'n':
                    # --- Guardar la venta en Excel ---
                    print(df_venta)
                    print(f'Total a pagar: {total}')
                    agregar_producto = False
                    resp_correcta = True

                    try:
                        libro = load_workbook(self.inventario)
                        hojas = libro.sheetnames

                    except FileNotFoundError:
                        hojas = []

                    # Si ya existe la hoja 'Ventas', anexar al final
                    if 'Ventas' in hojas:
                        df_ventas_dia = pd.read_excel(self.inventario, sheet_name='Ventas')

                        with pd.ExcelWriter(self.inventario,
                                            mode='a',
                                            if_sheet_exists='overlay'
                                            ) as writer:
                            df_venta.to_excel(writer,
                                              sheet_name='Ventas', 
                                              startrow=len(df_ventas_dia)+1, 
                                              index=False, 
                                              header=False
                                              )
                            
                    else:
                        # Si no existe la hoja 'Ventas', crear una nueva
                        with pd.ExcelWriter(self.inventario,
                                            mode='a',
                                            if_sheet_exists='new'
                                            ) as writer:
                            df_venta.to_excel(writer,
                                              sheet_name='Ventas', 
                                              index=False, 
                                              header=True
                                              )
                    
                    print('Venta registrada con exito')
                    print('Vuelve pronto!')
                    return df_venta

                else:
                    print('La opción ingresada no es valida')
                    resp_correcta = False

    # --- Metodo para actualizar el inventario despues de una venta
    def actualizar_inventario_venta(self, venta):
        '''
        Actualiza las existencias en el inventario después de registrar una
        venta.
         
        El método recorre los productos vendidos y descuenta sus cantidades
        correspondientes del inventario. Luego sobrescribe la hoja 'Inventario'
        del archivo Excel con los valores actualizados.
        
        Parámetros:
         venta (DataFrame): detalle de los productos vendidos, con columnas
                            ['Producto', 'Cantidad', 'Unidad', 'Precio', 
                            'Subtotal'].
        
        Retorna:
            None
        '''
        print('\nActualizando inventario...')
        
        # --- DataFrame con la información de los productos vendidos ---
        df_venta = venta
        
        # --- Leer el inventario actual desde Excel ---
        df_inventario = pd.read_excel(self.inventario, sheet_name='Inventario')

        # --- Restar las cantidades vendidas del inventario ---
        for index, row in df_venta.iterrows():
                for index2, row2 in df_inventario.iterrows():
                    if row['Producto'] == row2['Producto']:
                            # Descontar la cantidad vendida
                            row2['Cantidad'] = row2['Cantidad'] - row['Cantidad']

                            # Actualizar el valor en el DataFrame
                            df_inventario.at[index2, 'Cantidad'] = row2['Cantidad']
                            
        # --- Guardar los cambios actualizados en el archivo Excel ---
        with pd.ExcelWriter(self.inventario,
                            mode='a',
                            if_sheet_exists='overlay'
                            ) as writer:
             df_inventario.to_excel(writer,
                               sheet_name='Inventario', 
                               index=False, 
                               header=True
                               )

        print('Inventario actualizado con exito')

    # --- Método para generar un resumen de ventas por fecha ---
    def resumen_ventas(self):
        '''
        Genera un resumen de las ventas realizadas en una fecha específica.
        
        El método permite al usuario ingresar una fecha y muestra:
            - Los productos vendidos ese día.
            - La cantidad total vendida por producto.
            - El subtotal acumulado por producto.
            - El total general de ventas del día.
        
        Utiliza las hojas 'Ventas' e 'Inventario' del archivo Excel.
         
        Retorna:
            None
        '''
        # --- Leer los datos de ventas e inventario ---
        df_ventas = pd.read_excel(self.inventario, sheet_name='Ventas')
        df_inventario = pd.read_excel(self.inventario, sheet_name='Inventario')

        # --- Convertir la columna 'Fecha' a tipo date ---        
        df_ventas['Fecha'] = pd.to_datetime(df_ventas['Fecha']).dt.date
        
        # --- Solicitar al usuario la fecha de consulta ---
        fecha = input('Ingresa la fecha (aaaa-mm-dd): ')
        fecha = pd.to_datetime(fecha).date()
        
        # --- Inicializar variables acumuladoras ---
        total_dia = 0
        cant_prod = 0
        total_prod = 0
        fecha_usuario = True

        print(f'\nResumen de ventas del dia {fecha}:')
        
        # --- Recorrer inventario y ventas para consolidar datos ---
        for index, row in df_inventario.iterrows():
            for index1, row1 in df_ventas.iterrows(): 
                # Verificar coincidencia de producto y fecha
                if (row1['Fecha'] == fecha) and (row['Producto'] == row1['Producto']):
                    cant_prod += row1['Cantidad']
                    total_prod += row1['Subtotal']
                    total_dia += row1['Subtotal']
                    precio = row1['Precio']

            # --- Mostrar resumen por producto ---
            try:
                print(f'- {row['Producto']}: {cant_prod} {row['Unidad']} a {precio} c/u. Subtotal: ${total_prod}')

                cant_prod = 0
                total_prod = 0
                fecha_usuario = True 

            # Si no se encuentra la fecha (no hubo ventas ese día)
            except UnboundLocalError:
                    fecha_usuario = False
                            
        # --- Mostrar el total del día o mensaje si no hubo ventas ---
        if fecha_usuario == True:
            print(f'\nTotal de ventas del {fecha}: ${total_dia}')

        else:
            print('\nLa fecha solicitada no registra ventas')

    '''
    ===========================================================================
                             Fin de la clase Tienda
    ===========================================================================
    La clase tienda concentra la lógica principal del programa:

        - Gestión de inventario y precios.
        - Registro y actualización de ventas.
        - Reportes diarios de ventas.
    
    Cada método fue diseñado bajo el enfoque de Programación Orientada a
    Objetos (POO), con énfasis en legibilidad, estructura modular y facilidad
    de mantenimiento.
    '''

# --- Inicio del programa principal ---

# Archivo de Excel que contiene los datos del inventario
inv = 'Tienda.xlsx'

# Crear el objeto 'inventario' a partir de la clase tienda
inventario = tienda(inv)

# --- Menu principal ---
print('\nBienvenido a la tienda')

salir = False

# Bucle principal del programa
while not salir:
    print('\n¿Qué deseas hacer?')

    # Mostrar opciones del menú y leer la elección del usuario
    opcion = int(input('1 -> Revisar el inventario \n' \
                       '2 -> Actualizar el inventario \n' \
                       '3 -> Actualizar precios \n' \
                       '4 -> Agregar un producto al inventario \n' \
                       '5 -> Vender productos \n' \
                       '6 -> Resumen de ventas\n' \
                       '7 -> Salir \n' \
                       '\nIngresa el numero de la opcion deseada: '
                       ))

    # Validar opción ingresada
    if (opcion < 1) or (opcion > 7):
        print('La opción ingresada no es valida')
        print('Por favor ingresa un numero del 1 al 7')

    elif opcion == 1:
        # Mostrar el inventario actual
        inventario.revisar()

    elif opcion == 2:
        # Actualizar la cantidad de un producto existente
        act_producto = input('¿Que producto deseas actualizar? ')

        inventario.actualizar_inventario(act_producto)

    elif opcion == 3:
        # Modificar precios de productos
        inventario.actualizar_precios()

    elif opcion == 4:
        # Agregar un nuevo producto al inventario
        prod = input('¿Que producto deseas agregar al inventario? ')
        
        inventario = tienda(inv, prod)

        prod_en_inventario = inventario.leer_inventario()

        if prod_en_inventario:
            print(f'{prod} ya se encuentra en el inventario')

        else:
            inventario.agregar()

    elif opcion == 5:
        # Registrar una venta y actualizar el inventario
        prod_venta = input('\n¿Que producto deseas? ')

        venta = tienda(inv, prod_venta)
        ultima_venta = venta.vender()

        inventario.actualizar_inventario_venta(ultima_venta)

    elif opcion == 6:
        # Mostrar resumen de ventas
        inventario.resumen_ventas()

    elif opcion == 7:
        # Salir del programa
        print('Hasta pronto!')

        salir = True

'''
===============================================================================
                            Fin del programa principal
===============================================================================
Autor: Lucas O. Urbano Bedoya
Versión: 1.0
Lenguaje: Python
Librerías: Pandas, Openpyxl

Notas para mantenimiento:
- Verificar compatibilidad con nuevas versiones de Pandas y Openpyxl.
- Agregar validaciones adicionales para entradas del usuario.
- Implementar manejo de excepciones más detallado (ValueError, 
  KeyError, etc.).
- Optimizar la escritura en Excel para grandes volúmenes de datos.
- Explorar interfaz gráfica (Tkinter o Streamlit) para mejorar la 
  experiencia.
'''